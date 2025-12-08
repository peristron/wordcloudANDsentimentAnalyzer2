#  OPTIMIZED FOR DEPLOYMENT + LARGE FILES (750MB+) + DATA PREVIEW
#
import io
import re
import html
import gc
import time
import csv
import json
import string
from collections import Counter
from typing import Dict, List, Tuple, Iterable, Optional, Callable

import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
from io import BytesIO
from wordcloud import WordCloud, STOPWORDS
from matplotlib import font_manager
from itertools import pairwise
import openai

# --- GRAPH IMPORTS ---
import networkx as nx
from streamlit_agraph import agraph, Node, Edge, Config

# --- OPTIONAL IMPORTS (Handle missing libraries gracefully) ---
try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pypdf
except ImportError:
    pypdf = None

try:
    import nltk
    from nltk.sentiment.vader import SentimentIntensityAnalyzer
except ImportError:
    nltk = None
    SentimentIntensityAnalyzer = None

# precompiled patterns
HTML_TAG_RE = re.compile(r"<[^>]+>")
CHAT_ARTIFACT_RE = re.compile(
    r":\w+:"
    r"|\b(?:monday|tuesday|wednesday|thursday|friday|saturday|sunday|today|yesterday) at \d{1,2}:\d{2}\b"
    r"|\b\d+\s+repl(?:y|ies)\b"
    r"|\d{2}:\d{2}:\d{2}\.\d{3}\s+-->\s+\d{2}:\d{2}:\d{2}\.\d{3}"
    r"|\[[^\]]+\]",
    flags=re.IGNORECASE
)

# ---------------------------
# AUTH & SESSION UTILS
# ---------------------------

if 'total_cost' not in st.session_state: st.session_state['total_cost'] = 0.0
if 'total_tokens' not in st.session_state: st.session_state['total_tokens'] = 0
if 'authenticated' not in st.session_state: st.session_state['authenticated'] = False
if 'auth_error' not in st.session_state: st.session_state['auth_error'] = False
if 'ai_response' not in st.session_state: st.session_state['ai_response'] = ""

def perform_login():
    password = st.session_state.password_input
    correct_password = st.secrets.get("auth_password", "admin")
    if password == correct_password:
        st.session_state['authenticated'] = True
        st.session_state['auth_error'] = False
        st.session_state['password_input'] = "" 
    else:
        st.session_state['auth_error'] = True

def logout():
    st.session_state['authenticated'] = False
    st.session_state['ai_response'] = ""

# ---------------------------
# UTILITIES & SETUP
# ---------------------------

@st.cache_resource(show_spinner="Initializing sentiment analyzer...")
def setup_sentiment_analyzer():
    if nltk is None: return None
    try: nltk.data.find('sentiment/vader_lexicon.zip')
    except LookupError: nltk.download('vader_lexicon')
    return SentimentIntensityAnalyzer()

def prefer_index(options: List[str], preferred: List[str]) -> int:
    for name in preferred:
        if name in options: return options.index(name)
    return 0 if options else -1

@st.cache_data(show_spinner=False)
def list_system_fonts() -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for fe in font_manager.fontManager.ttflist:
        if fe.name not in mapping: mapping[fe.name] = fe.fname
    return dict(sorted(mapping.items(), key=lambda x: x[0].lower()))

def build_punct_translation(keep_hyphens: bool, keep_apostrophes: bool) -> dict:
    punct = string.punctuation
    if keep_hyphens: punct = punct.replace("-", "")
    if keep_apostrophes: punct = punct.replace("'", "")
    return str.maketrans("", "", punct)

def parse_user_stopwords(raw: str) -> Tuple[List[str], List[str]]:
    phrases, singles = [], []
    for item in [x.strip() for x in raw.split(",") if x.strip()]:
        if " " in item: phrases.append(item.lower())
        else: singles.append(item.lower())
    return phrases, singles

def default_prepositions() -> set:
    return {'about', 'above', 'across', 'after', 'against', 'along', 'among', 'around', 'at', 'before', 'behind', 'below', 'beneath', 'beside', 'between', 'beyond', 'but', 'by', 'concerning', 'despite', 'down', 'during', 'except', 'for', 'from', 'in', 'inside', 'into', 'like', 'near', 'of', 'off', 'on', 'onto', 'out', 'outside', 'over', 'past', 'regarding', 'since', 'through', 'throughout', 'to', 'toward', 'under', 'underneath', 'until', 'up', 'upon', 'with', 'within', 'without'}

def build_phrase_pattern(phrases: List[str]) -> Optional[re.Pattern]:
    if not phrases: return None
    escaped = [re.escape(p) for p in phrases if p]
    if not escaped: return None
    return re.compile(rf"\b(?:{'|'.join(escaped)})\b", flags=re.IGNORECASE)

def estimate_row_count_from_bytes(file_bytes: bytes) -> int:
    if not file_bytes: return 0
    n = file_bytes.count(b"\n")
    if not file_bytes.endswith(b"\n"): n += 1
    return n

def format_duration(seconds: float) -> str:
    seconds = int(seconds)
    h, r = divmod(seconds, 3600)
    m, s = divmod(r, 60)
    return f"{h:d}:{m:02d}:{s:02d}" if h > 0 else f"{m:d}:{s:02d}"

def make_unique_header(raw_names: List[Optional[str]]) -> List[str]:
    seen: Dict[str, int] = {}
    result: List[str] = []
    for i, nm in enumerate(raw_names):
        name = (str(nm).strip() if nm is not None else "")
        if not name: name = f"col_{i}"
        if name in seen:
            seen[name] += 1
            unique = f"{name}__{seen[name]}"
        else:
            seen[name] = 1
            unique = name
        result.append(unique)
    return result

# ---------------------------
# ROW READERS (Enhanced for Multi-Format)
# ---------------------------

def read_rows_raw_lines(file_bytes: bytes, encoding_choice: str = "auto") -> Iterable[str]:
    def _iter_with_encoding(enc: str):
        bio = io.BytesIO(file_bytes)
        with io.TextIOWrapper(bio, encoding=enc, errors="replace", newline=None) as wrapper:
            for line in wrapper: yield line.rstrip("\r\n")
    if encoding_choice == "latin-1": yield from _iter_with_encoding("latin-1")
    else: yield from _iter_with_encoding("utf-8")

def read_rows_vtt(file_bytes: bytes, encoding_choice: str = "auto") -> Iterable[str]:
    for line in read_rows_raw_lines(file_bytes, encoding_choice):
        line = line.strip()
        if not line or line == "WEBVTT" or "-->" in line or line.isdigit(): continue
        if ":" in line:
            parts = line.split(":", 1)
            if len(parts) > 1 and len(parts[0]) < 30 and " " in parts[0]:
                yield parts[1].strip()
                continue
        yield line

def read_rows_pdf(file_bytes: bytes) -> Iterable[str]:
    if pypdf is None: return
    bio = io.BytesIO(file_bytes)
    try:
        reader = pypdf.PdfReader(bio)
        for page in reader.pages:
            text = page.extract_text()
            if text:
                yield text
    except Exception:
        yield ""

def read_rows_json(file_bytes: bytes, selected_key: str = None) -> Iterable[str]:
    bio = io.BytesIO(file_bytes)
    # Try reading as line-delimited JSON (JSONL) first
    try:
        wrapper = io.TextIOWrapper(bio, encoding="utf-8", errors="replace")
        for line in wrapper:
            if not line.strip(): continue
            try:
                obj = json.loads(line)
                if selected_key and isinstance(obj, dict):
                    yield str(obj.get(selected_key, ""))
                elif isinstance(obj, str):
                    yield obj
                else:
                    yield str(obj)
            except json.JSONDecodeError:
                # If fail on line 1, might be standard JSON array
                bio.seek(0)
                data = json.load(wrapper)
                if isinstance(data, list):
                    for item in data:
                        if selected_key and isinstance(item, dict):
                            yield str(item.get(selected_key, ""))
                        else:
                            yield str(item)
                elif isinstance(data, dict):
                     # If it's a single dict, try to yield the key or just values
                     if selected_key: yield str(data.get(selected_key, ""))
                     else: yield str(data)
                break 
    except Exception:
        pass

# --- CSV / Excel Utils ---

def detect_csv_num_cols(file_bytes: bytes, encoding_choice: str = "auto", delimiter: str = ",") -> int:
    enc = "latin-1" if encoding_choice == "latin-1" else "utf-8"
    bio = io.BytesIO(file_bytes)
    with io.TextIOWrapper(bio, encoding=enc, errors="replace", newline="") as wrapper:
        rdr = csv.reader(wrapper, delimiter=delimiter)
        row = next(rdr, None)
        return len(row) if row is not None else 0

def get_csv_columns(file_bytes: bytes, encoding_choice: str = "auto", delimiter: str = ",", has_header: bool = True) -> List[str]:
    enc = "latin-1" if encoding_choice == "latin-1" else "utf-8"
    bio = io.BytesIO(file_bytes)
    with io.TextIOWrapper(bio, encoding=enc, errors="replace", newline="") as wrapper:
        rdr = csv.reader(wrapper, delimiter=delimiter)
        first = next(rdr, None)
        if first is None: return []
        return make_unique_header(first) if has_header else [f"col_{i}" for i in range(len(first))]

def get_csv_preview(file_bytes: bytes, encoding_choice: str = "auto", delimiter: str = ",", has_header: bool = True, rows: int = 5) -> pd.DataFrame:
    enc = "latin-1" if encoding_choice == "latin-1" else "utf-8"
    bio = io.BytesIO(file_bytes)
    try:
        df = pd.read_csv(bio, delimiter=delimiter, header=0 if has_header else None, nrows=rows, encoding=enc, on_bad_lines='skip')
        # If no header, rename cols
        if not has_header:
            df.columns = [f"col_{i}" for i in range(len(df.columns))]
        return df
    except:
        return pd.DataFrame()

def iter_csv_selected_columns(file_bytes: bytes, encoding_choice: str, delimiter: str, has_header: bool, selected_columns: List[str], join_with: str = " ", drop_empty: bool = True) -> Iterable[str]:
    enc = "latin-1" if encoding_choice == "latin-1" else "utf-8"
    bio = io.BytesIO(file_bytes)
    with io.TextIOWrapper(bio, encoding=enc, errors="replace", newline="") as wrapper:
        rdr = csv.reader(wrapper, delimiter=delimiter)
        first = next(rdr, None)
        if first is None: return
        
        # Determine mapping based on header setting
        if has_header:
            header = make_unique_header(first)
            name_to_idx = {n: i for i, n in enumerate(header)}
        else:
            # If no header, cols are col_0, col_1 etc.
            name_to_idx = {f"col_{i}": i for i in range(len(first))}
            
            # Process the very first row since it is data
            idxs = [name_to_idx[n] for n in selected_columns if n in name_to_idx]
            vals = [first[i] if i < len(first) else "" for i in idxs]
            if drop_empty: vals = [v for v in vals if v]
            yield join_with.join(str(v) for v in vals)

        idxs = [name_to_idx[n] for n in selected_columns if n in name_to_idx]
        for row in rdr:
            vals = [row[i] if i < len(row) else "" for i in idxs]
            if drop_empty: vals = [v for v in vals if v]
            yield join_with.join(str(v) for v in vals)

def get_excel_sheetnames(file_bytes: bytes) -> List[str]:
    if openpyxl is None: return []
    bio = io.BytesIO(file_bytes)
    wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
    sheets = list(wb.sheetnames)
    wb.close()
    return sheets

def get_excel_preview(file_bytes: bytes, sheet_name: str, has_header: bool = True, rows: int = 5) -> pd.DataFrame:
    if openpyxl is None: return pd.DataFrame()
    bio = io.BytesIO(file_bytes)
    try:
        # read specific sheet
        df = pd.read_excel(bio, sheet_name=sheet_name, header=0 if has_header else None, nrows=rows, engine='openpyxl')
        if not has_header:
            df.columns = [f"col_{i}" for i in range(len(df.columns))]
        return df
    except:
        return pd.DataFrame()

def get_excel_columns(file_bytes: bytes, sheet_name: str, has_header: bool = True) -> List[str]:
    # Use preview to get cols
    df = get_excel_preview(file_bytes, sheet_name, has_header, rows=1)
    if not df.empty:
        return list(df.columns)
    return []

def excel_estimate_rows(file_bytes: bytes, sheet_name: str, has_header: bool = True) -> int:
    if openpyxl is None: return 0
    bio = io.BytesIO(file_bytes)
    wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
    ws = wb[sheet_name]
    total = ws.max_row or 0
    wb.close()
    if has_header and total > 0: total -= 1
    return max(total, 0)

def iter_excel_selected_columns(file_bytes: bytes, sheet_name: str, has_header: bool, selected_columns: List[str], join_with: str = " ", drop_empty: bool = True) -> Iterable[str]:
    if openpyxl is None: return
    bio = io.BytesIO(file_bytes)
    wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    first = next(rows_iter, None)
    if first is None: wb.close(); return
    
    # Map Columns
    if has_header:
        header = make_unique_header(list(first))
        name_to_idx = {n: i for i, n in enumerate(header)}
        idxs = [name_to_idx[n] for n in selected_columns if n in name_to_idx]
    else:
        header = [f"col_{i}" for i in range(len(first))]
        name_to_idx = {n: i for i, n in enumerate(header)}
        idxs = [name_to_idx[n] for n in selected_columns if n in name_to_idx]
        
        # Yield first row as data
        vals = [first[i] if i < len(first) else "" for i in idxs]
        if drop_empty: vals = [v for v in vals if v]
        yield join_with.join("" if v is None else str(v) for v in vals)

    for row in rows_iter:
        vals = [row[i] if (row is not None and i < len(row)) else "" for i in idxs]
        if drop_empty: vals = [v for v in vals if v]
        yield join_with.join("" if v is None else str(v) for v in vals)
    wb.close()

# ---------------------------
# CORE PROCESSING
# ---------------------------

def is_url_token(tok: str) -> bool:
    t = tok.strip("()[]{}<>,.;:'\"!?").lower()
    if not t: return False
    return ("://" in t) or t.startswith("www.")

def process_rows_iter(
    rows_iter: Iterable[str],
    remove_chat_artifacts: bool, remove_html_tags: bool, unescape_entities: bool, remove_urls: bool,
    keep_hyphens: bool, keep_apostrophes: bool,
    user_phrase_stopwords: Tuple[str, ...], user_single_stopwords: Tuple[str, ...],
    add_preps: bool, drop_integers: bool, min_word_len: int,
    compute_bigrams: bool = False, progress_cb: Optional[Callable[[int], None]] = None, update_every: int = 2_000,
) -> Dict:
    start_time = time.perf_counter()
    stopwords = set(STOPWORDS)
    stopwords.update(user_single_stopwords)
    if add_preps: stopwords.update(default_prepositions())
    translate_map = build_punct_translation(keep_hyphens=keep_hyphens, keep_apostrophes=keep_apostrophes)
    phrase_pattern = build_phrase_pattern(list(user_phrase_stopwords))
    counts = Counter()
    bigram_counts = Counter() if compute_bigrams else None
    total_rows = 0
    _remove_chat, _remove_html, _unescape, _remove_urls = remove_chat_artifacts, remove_html_tags, unescape_entities, remove_urls
    _min_len, _drop_int, _stopwords = min_word_len, drop_integers, stopwords
    _is_url, _trans, _ppat = is_url_token, translate_map, phrase_pattern

    for line in rows_iter:
        total_rows += 1
        text = line if isinstance(line, str) else ("" if line is None else str(line))
        if _remove_chat: text = CHAT_ARTIFACT_RE.sub(" ", text)
        if _remove_html: text = HTML_TAG_RE.sub(" ", text)
        if _unescape:
            try: text = html.unescape(text)
            except MemoryError: pass
        text = text.lower()
        if _ppat: text = _ppat.sub(" ", text)
        filtered_tokens_line: List[str] = []
        for t in text.split():
            if _remove_urls and _is_url(t): continue
            t = t.translate(_trans)
            if not t or len(t) < _min_len or (_drop_int and t.isdigit()) or t in _stopwords: continue
            filtered_tokens_line.append(t)
        if filtered_tokens_line:
            counts.update(filtered_tokens_line)
            if compute_bigrams and len(filtered_tokens_line) > 1:
                bigram_counts.update(tuple(bg) for bg in pairwise(filtered_tokens_line))
        if progress_cb and (total_rows % update_every == 0): progress_cb(total_rows)

    if progress_cb: progress_cb(total_rows)
    elapsed = time.perf_counter() - start_time
    return {"counts": counts, "bigrams": bigram_counts or Counter(), "rows": total_rows, "elapsed": elapsed}

# ---------------------------
# STATS & ANALYTICS HELPERS
# ---------------------------

def calculate_text_stats(counts: Counter, total_rows: int) -> Dict:
    total_tokens = sum(counts.values())
    unique_tokens = len(counts)
    # Estimate average word length weighted by frequency
    avg_len = sum(len(word) * count for word, count in counts.items()) / total_tokens if total_tokens else 0
    
    return {
        "Total Rows": total_rows,
        "Total Tokens": total_tokens,
        "Unique Vocabulary": unique_tokens,
        "Avg Word Length": round(avg_len, 2),
        "Lexical Diversity": round(unique_tokens / total_tokens, 4) if total_tokens else 0
    }

# ---------------------------
# SENTIMENT & VIZ
# ---------------------------

@st.cache_data(show_spinner="Analyzing term sentiment...")
def get_sentiments(_analyzer, terms: Tuple[str, ...]) -> Dict[str, float]:
    if not _analyzer or not terms: return {}
    return {term: _analyzer.polarity_scores(term)['compound'] for term in terms}

def create_sentiment_color_func(sentiments: Dict[str, float], pos_color: str, neg_color: str, neu_color: str, pos_threshold: float, neg_threshold: float):
    def color_func(word, font_size, position, orientation, random_state=None, **kwargs):
        score = sentiments.get(word, 0.0)
        if score >= pos_threshold: return pos_color
        elif score <= neg_threshold: return neg_color
        else: return neu_color
    return color_func

def get_sentiment_category(score: float, pos_threshold: float, neg_threshold: float) -> str:
    if score >= pos_threshold: return "Positive"
    if score <= neg_threshold: return "Negative"
    return "Neutral"

def build_wordcloud_figure_from_counts(counts: Counter, max_words: int, width: int, height: int, bg_color: str, colormap: str, font_path: Optional[str], random_state: int, color_func: Optional[Callable] = None):
    limited = dict(counts.most_common(max_words))
    wc = WordCloud(width=width, height=height, background_color=bg_color, colormap=colormap, font_path=font_path, random_state=random_state, color_func=color_func, collocations=False, normalize_plurals=False).generate_from_frequencies(limited)
    fig_w, fig_h = max(6.0, width / 100.0), max(3.0, height / 100.0)
    fig, ax = plt.subplots(figsize=(fig_w, fig_h), dpi=100)
    ax.imshow(wc, interpolation="bilinear")
    ax.axis("off")
    plt.tight_layout()
    return fig, wc

def fig_to_png_bytes(fig: plt.Figure) -> BytesIO:
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", pad_inches=0.1)
    buf.seek(0)
    return buf

# ---------------------------
# AI LOGIC
# ---------------------------
def generate_ai_insights(counts: Counter, bigrams: Counter, config: dict):
    try:
        top_unigrams = [w for w, c in counts.most_common(100)]
        top_bigrams = [" ".join(bg) for bg, c in bigrams.most_common(30)] if bigrams else ["(Bigrams disabled)"]
        
        context = f"""
        Top 100 Unigrams: {', '.join(top_unigrams)}
        Top 30 Bigrams: {', '.join(top_bigrams)}
        """
        
        system_prompt = """You are a qualitative data analyst. 
        Analyze the provided word frequency lists (extracted from a text corpus) to identify likely themes, topics, and context.
        Do NOT just list the words. Interpret them.
        Format your response with markdown headers.
        1. Likely Subject Matter
        2. Key Themes/Topics
        3. Potential Anomalies or Noise
        """
        
        client = openai.OpenAI(api_key=config['api_key'], base_url=config['base_url'])
        response = client.chat.completions.create(
            model=config['model_name'],
            messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": context}]
        )
        
        content = response.choices[0].message.content
        if hasattr(response, 'usage') and response.usage:
            in_tok = response.usage.prompt_tokens
            out_tok = response.usage.completion_tokens
            cost = (in_tok * config['price_in'] / 1_000_000) + (out_tok * config['price_out'] / 1_000_000)
            st.session_state['total_tokens'] += (in_tok + out_tok)
            st.session_state['total_cost'] += cost
            
        return content
    except Exception as e:
        return f"AI Error: {str(e)}"

# ---------------------------
# MAIN APP
# ---------------------------

st.set_page_config(page_title="Word Cloud & Graph Analytics", layout="wide")
st.title("🧠 Multi-File Word Cloud & Graph Analyzer")

st.warning("""
**⚠️ Data Privacy & Security Notice**
When in doubt, pre-sanitize your data. This is a public app running on Streamlit Community Cloud. **Do not upload files containing sensitive, private, or proprietary information.** All data you upload is processed on public servers.
""")

analyzer = setup_sentiment_analyzer()

# --- SIDEBAR START ---
with st.sidebar:
    st.header("🔐 AI Setup")
    if st.session_state['authenticated']:
        st.success("AI Features Unlocked")
        
        with st.expander("🤖 Provider Settings", expanded=True):
            ai_provider = st.radio("Provider", ["xAI (Grok)", "OpenAI (GPT-4o)"])
            
            if "OpenAI" in ai_provider:
                api_key_name = "openai_api_key"
                base_url = None 
                model_name = st.selectbox("Model", ["gpt-4o", "gpt-4o-mini"])
                if "mini" in model_name:
                    price_in, price_out = 0.15, 0.60
                else:
                    price_in, price_out = 2.50, 10.00
            else:
                api_key_name = "xai_api_key"
                base_url = "https://api.x.ai/v1"
                model_options = {
                    "Grok 4.1 Fast (Reasoning) [Best Value]": "grok-4-1-fast-reasoning",
                    "Grok 4": "grok-4-0709",
                    "Grok 2 (Legacy)": "grok-2-1212"
                }
                choice = st.selectbox("Model", list(model_options.keys()))
                model_name = model_options[choice]
                
                if "fast" in model_name:
                    price_in, price_out = 0.20, 0.50
                elif "grok-4" in model_name:
                    price_in, price_out = 3.00, 15.00
                else:
                    price_in, price_out = 2.00, 10.00

            api_key = st.secrets.get(api_key_name)
            if not api_key: api_key = st.text_input(f"Enter {api_key_name}", type="password")
            
            ai_config = {
                'api_key': api_key,
                'base_url': base_url,
                'model_name': model_name,
                'price_in': price_in,
                'price_out': price_out
            }

        with st.expander("💰 Cost Estimator", expanded=False):
            c1, c2 = st.columns(2)
            c1.markdown(f"**Tokens:**\n{st.session_state['total_tokens']:,}")
            c2.markdown(f"**Cost:**\n`${st.session_state['total_cost']:.5f}`")
            if st.button("Reset Cost"):
                st.session_state['total_cost'] = 0.0
                st.session_state['total_tokens'] = 0
                st.rerun()
        
        if st.button("Logout"): logout(); st.rerun()
    else:
        with st.expander("Unlock AI Features", expanded=True):
            st.text_input("Password", type="password", key="password_input", on_change=perform_login)
            if st.session_state['auth_error']: st.error("Incorrect password.")

    st.divider()
    st.info("Performance Tip: Streaming allows files up to ~1GB (but no need to push it)")
    uploaded_files = st.file_uploader(
        "upload files (csv, xlsx, json, txt, vtt, pdf)",
        type=["csv", "xlsx", "xlsm", "vtt", "txt", "json", "pdf"],
        accept_multiple_files=True
    )

    st.markdown("### 🎨 appearance")
    bg_color = st.color_picker("background color", value="#ffffff")
    colormap = st.selectbox("colormap", options=["viridis", "plasma", "inferno", "magma", "cividis", "tab10", "tab20", "Dark2", "Set3", "rainbow", "cubehelix", "prism", "Blues", "Greens", "Oranges", "Reds", "Purples", "Greys"], index=0)
    max_words = st.slider("max words in word cloud", 50, 3000, 1000, 50)
    width = st.slider("image width (px)", 600, 2400, 1200, 100)
    height = st.slider("image height (px)", 300, 1400, 600, 50)
    random_state = st.number_input("random seed", 0, value=42, step=1)

    st.markdown("### 🔬 sentiment analysis")
    enable_sentiment = st.checkbox("enable sentiment analysis", value=False)
    if enable_sentiment and analyzer is None:
        st.error("NLTK not found.")
        enable_sentiment = False
    pos_threshold, neg_threshold, pos_color, neu_color, neg_color = 0.05, -0.05, '#2ca02c', '#808080', '#d62728'
    if enable_sentiment:
        c1, c2 = st.columns(2)
        with c1: pos_threshold = st.slider("pos threshold", 0.0, 1.0, 0.05, 0.01)
        with c2: neg_threshold = st.slider("neg threshold", -1.0, 0.0, -0.05, 0.01)
        c1, c2, c3 = st.columns(3)
        with c1: pos_color = st.color_picker("pos color", value=pos_color)
        with c2: neu_color = st.color_picker("neu color", value=neu_color)
        with c3: neg_color = st.color_picker("neg color", value=neg_color)

    st.markdown("### 🧹 cleaning")
    remove_chat_artifacts = st.checkbox("remove chat artifacts", value=True)
    remove_html_tags = st.checkbox("strip html tags", value=True)
    unescape_entities = st.checkbox("unescape html entities", value=True)
    remove_urls = st.checkbox("remove urls", value=True)
    keep_hyphens = st.checkbox("keep hyphens", value=False)
    keep_apostrophes = st.checkbox("keep apostrophes", value=False)

    st.markdown("### 🛑 stopwords")
    user_input = st.text_area("custom stopwords (comma-separated)", value="firstname.lastname, jane doe")
    user_phrase_stopwords, user_single_stopwords = parse_user_stopwords(user_input)
    add_preps = st.checkbox("remove prepositions", value=True)
    drop_integers = st.checkbox("remove integers", value=True)
    min_word_len = st.slider("min word length", 1, 10, 2)

    st.markdown("### 📊 tables & font")
    top_n = st.number_input("top terms count", 5, 10000, 20)
    font_map, font_names = list_system_fonts(), list(list_system_fonts().keys())
    preferred_defaults = ["cmtt10", "cmr10", "Arial", "DejaVu Sans", "Helvetica", "Verdana"]
    default_font_index = prefer_index(font_names, preferred_defaults)
    combined_font_name = st.selectbox("font for combined cloud", font_names or ["(default)"], max(default_font_index, 0))
    combined_font_path = font_map.get(combined_font_name) if font_names else None
    
    with st.expander("⚙️ performance options", expanded=True):
        encoding_choice = st.selectbox("file encoding", ["auto (utf-8)", "latin-1"])
        chunksize = st.number_input("csv chunk size", 1_000, 100_000, 10_000, 1_000)
        compute_bigrams = st.checkbox("compute bigrams / graph", value=True, help="Required for Network Graph features")

# ---------------------------
# MAIN PROCESSING LOOP
# ---------------------------
combined_counts, combined_bigrams, file_results = Counter(), Counter(), []
if uploaded_files:
    st.subheader("📄 Per-File Processing")
    overall_bar, overall_status = st.progress(0), st.empty()
    use_combined_option = "use combined font"
    total_files = len(uploaded_files)

    for idx, file in enumerate(uploaded_files):
        file_bytes, fname, lower = file.getvalue(), file.name, file.name.lower()
        
        # File type detection
        is_csv = lower.endswith(".csv")
        is_xlsx = lower.endswith((".xlsx", ".xlsm"))
        is_vtt = lower.endswith(".vtt")
        is_txt = lower.endswith(".txt")
        is_json = lower.endswith(".json")
        is_pdf = lower.endswith(".pdf")

        if font_names:
            per_file_font_choice = st.sidebar.selectbox(f"font for {fname}", [use_combined_option] + font_names, 0, key=f"font_{idx}")
            per_file_font_path = combined_font_path if per_file_font_choice == use_combined_option else font_map.get(per_file_font_choice)
        else: per_file_font_choice, per_file_font_path = "(default)", None
        
        # --- INPUT OPTIONS WITH DATA PREVIEW ---
        with st.expander(f"🧩 Input Options: {fname}", expanded=False):
            if is_vtt: st.info("VTT transcript detected.")
            elif is_pdf: st.info("PDF document detected. Processing page by page.")
            elif is_txt: st.info("Plain Text file detected.")
            
            elif is_csv:
                try: inferred_cols = detect_csv_num_cols(file_bytes, encoding_choice, delimiter=",")
                except Exception: inferred_cols = 1
                default_mode = "csv columns" if inferred_cols > 1 else "raw lines"
                read_mode = st.radio("read mode", ["raw lines", "csv columns"], index=0 if default_mode=="raw lines" else 1, key=f"csv_mode_{idx}")
                delim_choice = st.selectbox("delimiter", [",", "tab", ";", "|"], 0, key=f"csv_delim_{idx}")
                delimiter = {",": ",", "tab": "\t", ";": ";", "|": "|"}[delim_choice]
                has_header = st.checkbox("header row", value=True if inferred_cols > 1 else False, key=f"csv_header_{idx}")
                selected_cols, join_with = [], " "
                
                if read_mode == "csv columns":
                    # PREVIEW LOGIC
                    st.caption("🔍 Data Preview (First 5 Rows)")
                    df_prev = get_csv_preview(file_bytes, encoding_choice, delimiter, has_header)
                    st.dataframe(df_prev, use_container_width=True, height=150)
                    
                    if not df_prev.empty:
                        col_names = list(df_prev.columns)
                        selected_cols = st.multiselect("Select Text Columns", col_names, [col_names[0]], key=f"csv_cols_{idx}")
                        join_with = st.text_input("join with", " ", key=f"csv_join_{idx}")

            elif is_xlsx:
                if openpyxl:
                    sheets = get_excel_sheetnames(file_bytes)
                    sheet_name = st.selectbox("sheet", sheets or ["(none)"], 0, key=f"xlsx_sheet_{idx}")
                    has_header = st.checkbox("header row", True, key=f"xlsx_header_{idx}")
                    
                    # PREVIEW LOGIC
                    if sheet_name:
                        st.caption("🔍 Data Preview (First 5 Rows)")
                        df_prev = get_excel_preview(file_bytes, sheet_name, has_header)
                        st.dataframe(df_prev, use_container_width=True, height=150)
                        
                        if not df_prev.empty:
                            col_names = list(df_prev.columns)
                            selected_cols = st.multiselect("Select Text Columns", col_names, [col_names[0]], key=f"xlsx_cols_{idx}")
                            join_with = st.text_input("join with", " ", key=f"xlsx_join_{idx}")

            elif is_json:
                st.info("JSON/JSONL File. Reads as line-delimited or standard array.")
                json_key = st.text_input("Key to Extract (leave empty to read all text)", "", key=f"json_key_{idx}", help="If your JSON has objects like {'text': 'hello'}, enter 'text' here.")

        # --- PROCESSING CONTAINER ---
        container = st.container()
        with container:
            st.markdown(f"#### {fname}")
            per_file_bar, per_file_status = st.progress(0), st.empty()
        
        rows_iter, approx_rows = iter([]), 0
        
        # Setup Iterator based on type
        if is_vtt:
            rows_iter = read_rows_vtt(file_bytes, "latin-1" if encoding_choice == "latin-1" else "auto")
            approx_rows = estimate_row_count_from_bytes(file_bytes)
        elif is_pdf:
             rows_iter = read_rows_pdf(file_bytes)
             # Approx pages? Hard to know without parsing. Just dummy it or use len(reader.pages) if fast
             approx_rows = 0 
        elif is_txt:
            rows_iter = read_rows_raw_lines(file_bytes, "latin-1" if encoding_choice == "latin-1" else "auto")
            approx_rows = estimate_row_count_from_bytes(file_bytes)
        elif is_json:
            key_sel = locals().get('json_key', None)
            key_sel = key_sel if key_sel and key_sel.strip() else None
            rows_iter = read_rows_json(file_bytes, key_sel)
            approx_rows = estimate_row_count_from_bytes(file_bytes)
        elif is_csv:
            rmode = locals().get('read_mode', "raw lines")
            if rmode == "raw lines":
                rows_iter = read_rows_raw_lines(file_bytes, "latin-1" if encoding_choice == "latin-1" else "auto")
            else:
                rows_iter = iter_csv_selected_columns(file_bytes, "latin-1" if encoding_choice == "latin-1" else "auto", delimiter, has_header, selected_cols, join_with)
            approx_rows = estimate_row_count_from_bytes(file_bytes)
        elif is_xlsx and openpyxl:
            if sheet_name:
                rows_iter = iter_excel_selected_columns(file_bytes, sheet_name, has_header, selected_cols, join_with)
                approx_rows = excel_estimate_rows(file_bytes, sheet_name, has_header)
        else:
            rows_iter = read_rows_raw_lines(file_bytes, "latin-1" if encoding_choice == "latin-1" else "auto")
            approx_rows = estimate_row_count_from_bytes(file_bytes)
        
        # Update freq logic
        update_every = 500 if approx_rows <= 50000 else (2000 if approx_rows <= 500000 else 10000)
        start_wall = time.perf_counter()
        
        def make_progress_cb(total_hint: int):
            def _cb(done: int):
                elapsed = time.perf_counter() - start_wall
                if total_hint > 0:
                    per_file_bar.progress(min(99, int(done * 100 / total_hint)))
                    per_file_status.markdown(f"rows: {done:,}/{total_hint:,} • {format_duration(elapsed)}")
                else:
                    per_file_status.markdown(f"rows: {done:,} • {format_duration(elapsed)}")
            return _cb
        
        data = process_rows_iter(
            rows_iter, remove_chat_artifacts, remove_html_tags, unescape_entities, remove_urls,
            keep_hyphens, keep_apostrophes, tuple(user_phrase_stopwords), tuple(user_single_stopwords),
            add_preps, drop_integers, min_word_len, compute_bigrams, make_progress_cb(approx_rows), update_every
        )
        
        file_results.append({"rows": data["rows"]})

        per_file_bar.progress(100)
        per_file_status.markdown(f"✅ done in {format_duration(time.perf_counter() - start_wall)} • rows: {data['rows']:,}")
        combined_counts.update(data["counts"])
        if compute_bigrams: combined_bigrams.update(data["bigrams"])

        if data["counts"]:
            color_func = None
            if enable_sentiment:
                sentiments = get_sentiments(analyzer, tuple(data["counts"].keys()))
                color_func = create_sentiment_color_func(sentiments, pos_color, neg_color, neu_color, pos_threshold, neg_threshold)
            fig, _ = build_wordcloud_figure_from_counts(data["counts"], max_words, width, height, bg_color, colormap, per_file_font_path, random_state, color_func)
            col1, col2 = st.columns([3, 1])
            with col1: st.pyplot(fig, use_container_width=True)
            with col2: st.download_button(f"📥 download {fname} png", fig_to_png_bytes(fig), f"{fname}_wc.png", "image/png")
            plt.close(fig); gc.collect()
        else: st.warning(f"no tokens for {fname}.")
        overall_bar.progress(int(((idx + 1) / total_files) * 100))

# ---------------------------
# RESULTS & ANALYTICS
# ---------------------------
term_sentiments = {}
if enable_sentiment and combined_counts:
    term_sentiments = get_sentiments(analyzer, tuple(combined_counts.keys()))
    if compute_bigrams:
        bigram_phrases = tuple(" ".join(bg) for bg in combined_bigrams.keys())
        term_sentiments.update(get_sentiments(analyzer, bigram_phrases))

st.divider()
st.subheader("🖼️ Combined Word Cloud")
if combined_counts:
    try:
        c_color_func = None
        if enable_sentiment: c_color_func = create_sentiment_color_func(term_sentiments, pos_color, neg_color, neu_color, pos_threshold, neg_threshold)
        fig, _ = build_wordcloud_figure_from_counts(combined_counts, max_words, width, height, bg_color, colormap, combined_font_path, random_state, c_color_func)
        st.pyplot(fig, use_container_width=True)
        st.download_button("📥 download combined png", fig_to_png_bytes(fig), "combined_wc.png", "image/png")
        plt.close(fig); gc.collect()
    except MemoryError: st.error("memory error: reduce image size.")

# ---------------------------
# COMBINED ANALYTICS SECTION
# ---------------------------

if combined_counts:
    st.divider()
    
    total_processed_rows = sum(f['rows'] for f in file_results)
    text_stats = calculate_text_stats(combined_counts, total_processed_rows)

    show_graph = compute_bigrams and combined_bigrams and st.checkbox("🕸️ Show Network Graph & Advanced Analytics", value=True)
    
# You need to import math at the very top of your file
import math 

# ... [Inside the main app flow] ...

    if show_graph:
        st.subheader("🔗 Network Graph & Analytics")
        
        # 1. GRAPH CONFIGURATION
        with st.expander("🛠️ Graph Settings & Physics", expanded=False):
            c1, c2, c3 = st.columns(3)
            min_edge_weight = c1.slider("Min Link Frequency", 2, 100, 5, help="Filter out rare connections.")
            max_nodes_graph = c1.slider("Max Nodes", 10, 200, 50, help="Fewer nodes = cleaner graph.")
            
            # New Physics Controls
            repulsion = c2.slider("Node Spacing (Repulsion)", 100, 2000, 800, help="Push nodes further apart.")
            edge_len_val = c2.slider("Target Edge Length", 50, 500, 200, help="How long should the lines be?")
            
            physics_enabled = c3.checkbox("Enable Physics", True)
            directed_graph = c3.checkbox("Show Arrows (Directed)", False, help="Turn off to remove arrowheads and clean up the view.")

        # 2. BUILD GRAPH
        G = nx.DiGraph() if directed_graph else nx.Graph()
        
        # Filter Logic
        filtered_bigrams = {k: v for k, v in combined_bigrams.items() if v >= min_edge_weight}
        sorted_connections = sorted(filtered_bigrams.items(), key=lambda x: x[1], reverse=True)[:max_nodes_graph]
        
        if not sorted_connections:
            st.warning("No connections found. Try lowering 'Min Link Frequency'.")
        else:
            # Add edges
            for (source, target), weight in sorted_connections:
                G.add_edge(source, target, weight=weight)

            # 3. VISUALIZATION PREP
            nodes, edges = [], []
            
            # Calculate metrics for sizing
            try:
                deg_centrality = nx.degree_centrality(G)
            except:
                deg_centrality = {n: 1 for n in G.nodes()}
            
            # Determine max weight for relative scaling
            max_weight = max([w for s, t, w in sorted_connections]) if sorted_connections else 1

            for node_id in G.nodes():
                # Dynamic node sizing
                size = 15 + (deg_centrality.get(node_id, 0) * 40)
                
                # Dynamic coloring
                node_color = neu_color
                if enable_sentiment:
                    score = term_sentiments.get(node_id, 0)
                    if score >= pos_threshold: node_color = pos_color
                    elif score <= neg_threshold: node_color = neg_color
                
                nodes.append(Node(
                    id=node_id,
                    label=node_id,
                    size=size,
                    color=node_color,
                    # Tooltip shows detailed stats
                    title=f"Term: {node_id}\nFreq: {combined_counts.get(node_id, 0)}"
                ))

            for (source, target), weight in sorted_connections:
                # Logarithmic scaling: Prevents massive lines for high-frequency common words
                # Formula: Base Thickness + log(weight)
                scaled_width = 1.0 + math.log(weight) * 0.5
                
                edges.append(Edge(
                    source=source,
                    target=target,
                    width=scaled_width,
                    color="#cccccc",
                    # type="curvedCW" # Optional: makes lines curved
                ))

            # 4. PHYSICS CONFIGURATION (Fixes the "Blob" issue)
            config = Config(
                width=900, 
                height=600, 
                directed=directed_graph,
                physics=physics_enabled, 
                hierarchy=False,
                # ForceAtlas2Based is often better for text networks than the default BarnsHut
                physicsSettings={
                    "solver": "forceAtlas2Based",
                    "forceAtlas2Based": {
                        "gravitationalConstant": -abs(repulsion), # Negative pushes nodes apart
                        "springLength": edge_len_val,
                        "springConstant": 0.08,
                        "damping": 0.4
                    }
                }
            )
            
            # Render
            agraph(nodes=nodes, edges=edges, config=config)

            # 5. TABBED ANALYTICS
            st.markdown("### 📊 Graph Analytics")
            tab1, tab2, tab3, tab4, tab5 = st.tabs(["Basic Stats", "Degree Stats", "Centrality Measures", "Top Nodes", "Text Stats"])
            
            with tab1:
                col_b1, col_b2, col_b3 = st.columns(3)
                col_b1.metric("Nodes", G.number_of_nodes())
                col_b2.metric("Edges", G.number_of_edges())
                try: col_b3.metric("Density", f"{nx.density(G):.4f}")
                except: pass

            with tab2:
                degrees = [val for (node, val) in G.degree()]
                if degrees:
                    st.metric("Avg Connections", f"{sum(degrees)/len(degrees):.2f}")
                    st.bar_chart(pd.Series(degrees).value_counts().sort_index(), use_container_width=True)

            with tab3:
                try:
                    # Scipy is required here
                    dc = nx.degree_centrality(G)
                    bc = nx.betweenness_centrality(G, weight='weight')
                    cc = nx.closeness_centrality(G)
                    # PageRank needs scipy
                    pr = nx.pagerank(G, weight='weight')
                    
                    data = [{"Node": n, "Degree": dc[n], "Betweenness": bc[n], "Closeness": cc[n], "PageRank": pr[n]} for n in G.nodes()]
                    st.dataframe(pd.DataFrame(data).set_index("Node").sort_values("PageRank", ascending=False).style.background_gradient(cmap="Blues"), use_container_width=True)
                except ImportError:
                    st.error("dependency missing: Please install 'scipy' to see Advanced Centrality.")
                except Exception as e:
                    st.error(f"Calculation Error: {e}")

            with tab4:
                # Simple weight sum
                nw = {n: 0 for n in G.nodes()}
                for u, v, d in G.edges(data=True):
                    w = d.get('weight', 1)
                    nw[u] += w
                    nw[v] += w
                st.dataframe(pd.DataFrame(list(nw.items()), columns=["Node", "Total Weight"]).sort_values("Total Weight", ascending=False).head(50), use_container_width=True)

            with tab5:
                # Text Stats from earlier calculation
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Tokens", f"{text_stats['Total Tokens']:,}")
                c2.metric("Vocab", f"{text_stats['Unique Vocabulary']:,}")
                c3.metric("Diversity", f"{text_stats['Lexical Diversity']}")
                c4.metric("Avg Len", f"{text_stats['Avg Word Length']}")

else: 
    st.info("upload files to start.")


# --- AI ANALYSIS SECTION ---
if combined_counts and st.session_state['authenticated']:
    st.divider()
    st.subheader("🤖 AI Theme Detection")
    st.caption("Send the top 100 terms to the AI to detect likely topics and anomalies.")
    if st.button("✨ Analyze Themes with AI", type="primary"):
        with st.status("Analyzing top terms...", expanded=True) as status:
            response = generate_ai_insights(combined_counts, combined_bigrams if compute_bigrams else None, ai_config)
            st.session_state['ai_response'] = response
            status.update(label="Analysis Complete", state="complete", expanded=False)

if st.session_state['ai_response']:
    st.markdown("### 📋 AI Insights")
    st.markdown(st.session_state['ai_response'])
    st.divider()

# --- TABLES ---
if combined_counts:
    st.divider()
    st.subheader(f"📊 Frequency Tables (Top {top_n})")
    most_common = combined_counts.most_common(top_n)
    data = [[w, f] + ([term_sentiments.get(w,0), get_sentiment_category(term_sentiments.get(w,0), pos_threshold, neg_threshold)] if enable_sentiment else []) for w, f in most_common]
    cols = ["word", "count"] + (["sentiment", "category"] if enable_sentiment else [])
    st.dataframe(pd.DataFrame(data, columns=cols), use_container_width=True)
    
    if compute_bigrams and combined_bigrams:
        st.write("Bigrams")
        top_bg = combined_bigrams.most_common(top_n)
        bg_data = [[" ".join(bg), f] + ([term_sentiments.get(" ".join(bg),0), get_sentiment_category(term_sentiments.get(" ".join(bg),0), pos_threshold, neg_threshold)] if enable_sentiment else []) for bg, f in top_bg]
        bg_cols = ["bigram", "count"] + (["sentiment", "category"] if enable_sentiment else [])
        st.dataframe(pd.DataFrame(bg_data, columns=bg_cols), use_container_width=True)
